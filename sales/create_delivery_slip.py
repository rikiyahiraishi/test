from lib2to3.pgen2.pgen import DFAState
from .models import Customer, Order

import pandas as pd
import openpyxl
import os
import re
import datetime
import subprocess
import shutil
from . import log_out

##########################################################################
file_path = "C:/Users/rikiya/Desktop/web_app/config/_format/納品書_小松.xlsx"
log_dir = "C:/Users/rikiya/Desktop/web_app/config/_log"
# 納品書シートのPDF出力
delivery_pdf_dir = "C:/Users/rikiya/Desktop"
##########################################################################

def nohin_write(wb, df, delivery_number, post_code, address, company_name, today, logging):
    try:
        status = True
        # 納品書シートをアクティブシートに設定
        active_sheet = wb['納品書']

        # 宛先会社情報の入力
        active_sheet['A6'] = "〒 " + post_code
        active_sheet['A7'] = "    " + address
        active_sheet['A8'] = company_name
        # 宛先会社情報の入力(控え用)
        active_sheet['A33'] = "〒 " + post_code
        active_sheet['A34'] = "    " + address
        active_sheet['A35'] = company_name

        # 納品番号、納品日の書き込み
        active_sheet['J6'] = delivery_number
        active_sheet['J7'] = today
        # 納品番号、納品日の書き込み(控え用)
        active_sheet['J33'] = delivery_number
        active_sheet['J34'] = today

        # 前回のデータ削除
        for row in active_sheet.iter_rows(min_row=16, min_col=1, max_row=25, max_col=1):
            for cell in row:
                cell.value = None
        for row in active_sheet.iter_rows(min_row=16, min_col=3, max_row=25, max_col=7):
            for cell in row:
                cell.value = None
        # 前回のデータ削除(控え用)
        for row in active_sheet.iter_rows(min_row=43, min_col=1, max_row=52, max_col=1):
            for cell in row:
                cell.value = None
        for row in active_sheet.iter_rows(min_row=43, min_col=3, max_row=52, max_col=7):
            for cell in row:
                cell.value = None

        # データの成型
        row_list = []
        # [0]:商品名、[1]:数量、[2]:単位、[3]:単価、[4]:金額、[5]:納品日
        for index, row in df.iterrows():
            new_row = [row[0], row[1], row[2], row[3], row[4], row[5]]
            row_list.append(new_row)
        # 「以下余白」の追加(10件未満の場合)
        if int(len(row_list)) < 10:
            row_list.append(["以下余白", "" , "" , "" , "", "", ""])

        # 納品商品データの書き込み
        col_list = ['A', 'C', 'D', 'E', 'F', 'G']
        for i in range(len(row_list)):
            row = row_list[i]
            for t in range(len(col_list)):
                STR = '{0}{1}'.format(col_list[t], 16+i)
                active_sheet[STR] = row[t]
                
        # 納品商品データの書き込み(控え用)
        for i in range(len(row_list)):
            row = row_list[i]
            for t in range(len(col_list)):
                STR = '{0}{1}'.format(col_list[t], 43+i)
                active_sheet[STR] = row[t]

    except Exception as err:
        logging.exception(err)

        logging.info("Process Failed ---make_delivery_slip---")
        status = False

    return active_sheet, status



def list_to_df(pk_list):
    df_list = list()
    a = pk_list
    for pk in pk_list:
        object = Order.objects.get(pk=pk)
        product_name = object.product_name
        volume = object.volume
        unit_price = object.unit_price
        unit = object.unit
        total_price = object.total_price
        delivery_date = object.delivery_date
        if not delivery_date:
            delivery_date = '指定なし'
        df_list.append([product_name, volume, unit_price, unit, total_price, delivery_date])

    df = pd.DataFrame(df_list)

    return df



def create_delivery_slip(pk_list, delivery_number, login_user):
    try:
        # ログモジュールの初期化
        logging = log_out.setup_logging(log_dir)
        logging.info("Process Start ---make_deliveryslip---")

        # openpyxlでのエクセルファイルの読み込み
        wb = openpyxl.load_workbook(file_path)

        # シートの読み込み
        sheet_obj = wb.sheetnames
        # シート名のリスト化
        sheet_list = []
        for i in range(len(sheet_obj)):
            sheet_list.append(sheet_obj[i])

        # 納品書シートの存在確認
        if not "納品書" in sheet_list:
            logging.error("「納品書」シートが存在しません。")
            logging.info("Process End ---make_deliveryslip---")
            return '', False
        
        # pk_listのdf化
        df = list_to_df(pk_list)

        # pk_listから顧客情報を取得
        customer_name = Order.objects.get(pk=pk_list[0]).customer_name
        customer_info = Customer.objects.get(customer_name=customer_name, user=login_user)
        address = customer_info.address
        post_code = customer_info.post_code

        # 納品書シートデータの書き込み
        if not delivery_number:
            delivery_number = "-"
        today = datetime.date.today()
        active_sheet, status = nohin_write(wb, df, delivery_number, post_code, address, customer_name, today, logging)

        # 書き込みに失敗したら処理を終了する
        if not status:
            message = '納品書の書き込み処理に失敗しました。ログ内容を管理者に連絡してください。'
            return '', False, message

        # PDFファイルパスの生成
        # ディレクトリが存在しない場合は、ディレクトリを作成
        delivery_dir = "{0}/{1}/{2}/{3}/{4}/{5}/{6}/".format(delivery_pdf_dir, "web_app", "UserID", login_user, customer_name, "納品書", today.year)
        if not os.path.exists(delivery_dir):
            os.makedirs(delivery_dir, exist_ok=False)

        # ファイル一覧を取得
        files_list = os.listdir(delivery_dir)
        # ディレクトリを除き、ファイルのみを一覧に編集
        files_list = [s for s in files_list if re.match('\d{8}_納品書_.+_\d+\.pdf', s)] 

        # ファイルが存在しない場合の処理
        date = today.strftime('%Y%m%d')
        if not files_list:
            new_file_name = "{0}_納品書_{1}_0001.pdf".format("".join(date), customer_name)
        else:
            # 連番の中で最新の番号を取得し、次の番号を設定する
            files_list.sort()
            last_file = files_list[-1]
            file_num = last_file.split("_")[-1]
            file_num = file_num.split(".")[-2]
            next_num = int(file_num) + 1

            diff = len(file_num) - len(str(next_num))

            zero_num = ""
            for i in range(diff):
                zero_num = zero_num + "0"

            new_num = zero_num + str(next_num)
            pre_file = "{0}_納品書_{1}".format("".join(date), customer_name)
            new_file_name = "{0}_{1}.pdf".format(pre_file, new_num)

        # bookの保存
        wb.save(file_path)

        #ファイル名の生成
        new_file_path = os.path.join(delivery_dir, new_file_name)
        # PDFファイルの出力
        subprocess.run(['C:\Program Files\LibreOffice\program\soffice.com', '--headless', '--convert-to', 'pdf', '--outdir', r'{0}'.format(delivery_pdf_dir), r'{0}'.format(file_path)])
        # PDFファイルの移動
        pdf_path = os.path.join(delivery_pdf_dir, '{0}.pdf'.format(os.path.splitext(os.path.basename(file_path))[0]))
        shutil.move(pdf_path, new_file_path)

    except Exception as err:
        logging.exception(err)
        logging.info("Process End ---make_deliveryslip---")

        # 正常に終了しなかったことを通知
        message = '納品書の作成処理に失敗しました。ログ内容を管理者に連絡してください。'
        return '', False, message

    return new_file_path, True, '納品書作成処理が正常に完了しました。'