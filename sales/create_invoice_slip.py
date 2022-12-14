from lib2to3.pgen2.pgen import DFAState
from .models import Products, Customer, Order

import pandas as pd
import openpyxl
import sys
import os
import re
import datetime

from openpyxl.styles.fonts import Font
from openpyxl.styles import Alignment
from openpyxl.styles.borders import Border, Side
import subprocess
import shutil
from . import log_out

##########################################################################
file_path = "C:/Users/rikiya/Desktop/web_app/config/_format/請求書_小松.xlsx"
log_dir = "C:/Users/rikiya/Desktop/web_app/config/_log"
# 請求書シートのPDF出力
invoice_pdf_dir = "C:/Users/rikiya/Desktop"
##########################################################################

def invoice_write(wb, df, invoice_date, invoice_number, invoice_period, post_code, address, company_name, logging):
    try:
        status = True
        # 請求書シートをアクティブシートに設定
        active_sheet = wb['請求書']

        # 宛先会社情報の入力
        active_sheet['A5'] = "〒 " + post_code
        active_sheet['A6'] = "    " + address
        active_sheet['A7'] = company_name

        # 請求番号、請求日の書き込み
        active_sheet['I5'] = invoice_number
        active_sheet['I6'] = invoice_date

        # 請求期間の書き込み
        active_sheet['B10'] = invoice_period

        # 前回のデータ削除
        tf = del_past_data(active_sheet, logging)
        if not tf:
            # エラーを出力する
            logging.error("前回データを削除できなったので、終了します。")
            logging.info("Process End ---make_invoice---")
            sys.exit()

        # データの成型
        row_list = []
        # [0]:受注日時、[1]:受注番号、[2]:納品日、[3]:商品名、[4]:数量、[5]:単位、[6]:単価、[7]:金額
        for index, row in df.iterrows():
            new_row = [row[0], row[1], row[2], row[3], row[4], row[5], row[6]]
            row_list.append(new_row)
        # 「以下余白」の追加
        row_list.append(["以下余白", "" , "" , "" , "", "", ""])

        # 追加する行数
        add_row = 0
        col_list = ['A', 'B', 'F', 'G', 'H', 'I', 'J']
        # row_listの行数をカウントし、データ数が25件より多い場合は、その分だけ行を追加する
        if len(row_list) > 25:
            add_row = len(row_list)-25
            for i in range(len(row_list)-25):
                active_sheet.insert_rows(18)
                
                # 追加行のフォント変更処理
                for col in col_list:
                    STR = '{0}18'.format(col)
                    active_sheet[STR].font = Font(name='メイリオ', size=24)
                    active_sheet[STR].alignment = Alignment(horizontal='center', vertical='center')

                # 表示フォーマットの変更（日付）    
                active_sheet['A18'].number_format = "yyyy/mm/dd"
                active_sheet['I18'].number_format = "yyyy/mm/dd"

                # 罫線の設定
                side = Side(style='medium', color='AEAAAA')
                active_sheet['A18'].border = Border(left=side)
                active_sheet['J18'].border = Border(right=side)

                # 42行目以降の商品名欄の結合と中央揃え
                STR = 'B{0}:E{0}'.format(42+i)
                active_sheet.merge_cells(range_string=STR)

        # 納品商品データの書き込み
        for i in range(len(row_list)):
            row = row_list[i]
            for t in range(len(col_list)):
                STR = '{0}{1}'.format(col_list[t], 17+i)
                active_sheet[STR] = row[t]

        # 合計処理の記入
        active_sheet['J{0}'.format(44+add_row)] = '=SUM(J17:J{0})'.format(41+add_row)
        active_sheet['J{0}'.format(46+add_row)] = '=ROUNDDOWN(J{0}*J{1},0)'.format(44+add_row, 45+add_row)
        active_sheet['J{0}'.format(47+add_row)] = '=J{0}+J{1}'.format(44+add_row, 46+add_row)

        active_sheet['B11'] = '=J{0}'.format(47+add_row)

    except Exception as err:
        logging.exception(err)
        logging.info("Process Failed ---make_invoiceslip---")
        status = False

    return active_sheet, status



def del_past_data(active_sheet, logging):
    status = True
    try:
        # A列の最終行を取得
        maxRow = active_sheet.max_row + 1
        maxClm = active_sheet.max_column + 1

        for j in range(1, maxClm):
            # 列の指定
            if j == 1:
                for i in reversed(range(1, maxRow)):
                    if active_sheet.cell(row=i, column=j).value != None:
                        last_row = i
                        break

        # 最終行が17以下ならデータが存在しないので、17に設定
        if last_row <= 17:
            last_row = 17

        # 前回のデータ削除
        for row in active_sheet.iter_rows(min_row=17, min_col=1, max_row=last_row, max_col=2):
            for cell in row:
                cell.value = None
        for row in active_sheet.iter_rows(min_row=17, min_col=6, max_row=last_row, max_col=10):
            for cell in row:
                cell.value = None

        # 複数ぺージになっていた場合に、単一ページに戻す（削除）
        if last_row > 41:
            for i in range(last_row - 41):
                active_sheet.delete_rows(18)

    except Exception as err:
        logging.exception(err)

        # 正常に終了しなかったことを表示
        status = False

        logging.info("Process End ---make_invoice---")
        sys.exit()


    return status



def list_to_df(pk_list):
    df_list = list()
    a = pk_list
    for pk in pk_list:
        object = Order.objects.get(pk=pk)
        order_datetime = object.order_datetime
        order_date = order_datetime.date()
        delivery_date = object.delivery_date
        product_name = object.product_name
        volume = object.volume
        unit_price = object.unit_price
        unit = object.unit
        total_price = object.total_price

        df_list.append([order_date, product_name, volume, unit_price, unit, delivery_date, total_price])

    df = pd.DataFrame(df_list)

    return df



def create_invoice_slip(pk_list, invoice_number, invoice_period, login_user):
    try:
        # ログモジュールの初期化
        logging = log_out.setup_logging(log_dir)
        logging.info("Process Start ---make_invoiceslip---")

        # openpyxlでのエクセルファイルの読み込み
        wb = openpyxl.load_workbook(file_path)

        # シートの読み込み
        sheet_obj = wb.sheetnames
        # シート名のリスト化
        sheet_list = []
        for i in range(len(sheet_obj)):
            sheet_list.append(sheet_obj[i])

        # 納品書シートの存在確認
        if not "請求書" in sheet_list:
            logging.error("「請求書」シートが存在しません。")
            logging.info("Process End ---make_invoiceslip---")
            return '', False
        
        # pk_listのdf化
        df = list_to_df(pk_list)

        # pk_listから顧客情報を取得
        customer_name = Order.objects.get(pk=pk_list[0]).customer_name
        customer_info = Customer.objects.get(customer_name=customer_name, user=login_user)
        address = customer_info.address
        post_code = customer_info.post_code

        # 納品書シートデータの書き込み
        if not invoice_number:
            invoice_number = "-"
        today = datetime.date.today()
        active_sht, status = invoice_write(wb, df, today, invoice_number, invoice_period, post_code, address, customer_name, logging)

        # 書き込みに失敗したら処理を終了する
        if not status:
            message = '請求書の書き込み処理に失敗しました。ログ内容を管理者に連絡してください。'
            return '', False, message

        # PDFファイルパスの生成
        # ディレクトリが存在しない場合は、ディレクトリを作成
        invoice_dir = "{0}/{1}/{2}/{3}/{4}/{5}/{6}/".format(invoice_pdf_dir, "web_app", "UserID", login_user, customer_name, "請求書", today.year)
        if not os.path.exists(invoice_dir):
            os.makedirs(invoice_dir, exist_ok=False)

        # ファイル一覧を取得
        files_list = os.listdir(invoice_dir)
        # ディレクトリを除き、ファイルのみを一覧に編集
        files_list = [s for s in files_list if re.match('\d{8}_請求書_.+_\d+\.pdf', s)] 

        # ファイルが存在しない場合の処理
        date = today.strftime('%Y%m%d')
        if not files_list:
            new_file_name = "{0}_請求書_{1}_0001.pdf".format("".join(date), customer_name)
        else:
            # 連番の中で最新の番号を取得し、次の番号を設定する
            files_list.sort()
            last_file = files_list[-1]
            file_num = last_file.split("_")[-1]
            file_num = file_num.split(".")[-2]
            next_num = int(file_num) + 1

            diff = len(file_num) - len(str(next_num))

            zero_num = ""
            for i in range(diff):
                zero_num = zero_num + "0"

            new_num = zero_num + str(next_num)
            pre_file = "{0}_請求書_{1}".format("".join(date), customer_name)
            new_file_name = "{0}_{1}.pdf".format(pre_file, new_num)

        # bookの保存
        wb.save(file_path)

        #ファイル名の生成
        new_file_path = os.path.join(invoice_dir, new_file_name)
        # PDFファイルの出力
        subprocess.run(['C:\Program Files\LibreOffice\program\soffice.com', '--headless', '--convert-to', 'pdf', '--outdir', r'{0}'.format(invoice_pdf_dir), r'{0}'.format(file_path)])
        # PDFファイルの移動
        pdf_path = os.path.join(invoice_pdf_dir, '{0}.pdf'.format(os.path.splitext(os.path.basename(file_path))[0]))
        shutil.move(pdf_path, new_file_path)
    
    except Exception as err:
        logging.exception(err)
        logging.info("Process End ---make_invoiceslip---")

        # 正常に終了しなかったことを通知
        message = '請求書の作成処理に失敗しました。ログ内容を管理者に連絡してください。'
        return '', False, message

    return new_file_path, True, '請求書作成処理が正常に完了しました。'